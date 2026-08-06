# -*- coding: utf-8 -*-
"""
qa-design.json 검증 + QA_TEMPLATE.xlsx 에 값 채워 산출물 생성.

에이전트는 JSON만 만들고, 서식·수식·집계는 이 스크립트가 처리한다.
"""
import json, shutil, sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl 필요: pip install openpyxl")

TEMPLATE = "templates/QA_TEMPLATE_v2.xlsx"

# _TC 시트 열 매핑 (템플릿 9행 헤더 기준)
TC_COL = {"no": 2, "major": 3, "middle": 4, "minor": 5,
          "condition": 6, "order": 7, "expected": 8,
          "jira": 17, "priority": 18, "comment": 19}
TC_RESULT_COLS = range(9, 17)   # I~P 플랫폼별 결과
TC_START_ROW = 11

# Coverage 시트
COV_START, COV_MAX = 6, 21          # 1. 요구사항별 검증항목
EXC_START, EXC_MAX = 27, 31         # 2. 예외케이스
RISK_START, RISK_MAX = 39, 41       # 3. 리스크

VALID_PRIORITY = {"High", "Medium", "Low"}
PLACEHOLDER = ["확정 필요", "확인 필요", "미정", "TBD", "추후", "협의 필요", "정상적으로", "적절히"]


def validate(d):
    errors, warnings = [], []
    tcs = d.get("testcases", [])
    cov = d.get("coverage", [])

    if not tcs:
        errors.append("testcases 가 비어있음")
    if not cov:
        warnings.append("coverage 가 비어있음")

    seen = set()
    for i, t in enumerate(tcs):
        tag = t.get("no", f"[{i}번째]")
        for f in ("no", "major", "middle", "minor", "condition", "expected", "priority"):
            if not str(t.get(f, "")).strip():
                errors.append(f"{tag}: '{f}' 가 비어있음")

        if t.get("no") in seen:
            errors.append(f"{tag}: TC 번호 중복")
        seen.add(t.get("no"))

        if t.get("priority") not in VALID_PRIORITY:
            errors.append(f"{tag}: Priority '{t.get('priority')}' 는 High/Medium/Low 아님")

        # 기대결과가 판정 가능한가
        exp = str(t.get("expected", ""))
        for p in PLACEHOLDER:
            if p in exp:
                errors.append(f"{tag}: 기대결과에 모호한 표현('{p}'). 판정 가능한 문장으로 재작성 필요")
                break

        # 실행 순서가 번호 매겨진 절차인가
        cond = str(t.get("condition", ""))
        if cond and not cond.lstrip().startswith("1."):
            warnings.append(f"{tag}: 테스트 조건이 '1.' 로 시작하지 않음 (번호 매긴 절차 권장)")

    # 커버리지 ↔ TC 상호 참조
    covered = set()
    for c in cov:
        for n in c.get("tc_nos", []):
            covered.add(n)
            if n not in seen:
                errors.append(f"Coverage '{c.get('objective', '')}': 존재하지 않는 TC 번호 {n} 참조")

    orphan = seen - covered
    if orphan:
        warnings.append(f"어느 커버리지 항목에도 매핑되지 않은 TC {len(orphan)}건: {', '.join(sorted(orphan))}")

    return errors, warnings


def fill(d, out_path):
    shutil.copy(TEMPLATE, out_path)
    wb = openpyxl.load_workbook(out_path)

    # ---------- _TC ----------
    tc = wb["_TC"]
    tcs = d["testcases"]
    for i, t in enumerate(tcs):
        r = TC_START_ROW + i
        for key, col in TC_COL.items():
            if key == "jira":
                continue
            tc.cell(row=r, column=col).value = t.get(key, "")
        for c in TC_RESULT_COLS:
            tc.cell(row=r, column=c).value = "N/T"

    total = len(tcs)
    tc.cell(row=3, column=3).value = total     # Total Case
    tc.cell(row=4, column=3).value = 0         # Pass
    tc.cell(row=5, column=3).value = 0         # Fail
    tc.cell(row=6, column=3).value = 0         # N/A
    tc.cell(row=7, column=3).value = total     # N/T

    # ---------- Coverage ----------
    cov = wb["Coverage"]
    cov.cell(row=2, column=2).value = f"[{d.get('screen_id', '')}] {d.get('screen', '')} 테스트 커버리지"

    items = d.get("coverage", [])[:COV_MAX - COV_START + 1]
    for i, c in enumerate(items):
        r = COV_START + i
        cov.cell(row=r, column=2).value = c.get("spec_no", "")
        cov.cell(row=r, column=3).value = c.get("category", "")
        cov.cell(row=r, column=4).value = c.get("technique", "")
        cov.cell(row=r, column=5).value = c.get("objective", "")
        cov.cell(row=r, column=6).value = c.get("checklist", "")
        cov.cell(row=r, column=7).value = c.get("expected", "")
        cov.cell(row=r, column=8).value = ", ".join(c.get("tc_nos", []))
        cov.cell(row=r, column=9).value = len(c.get("tc_nos", []))
        cov.cell(row=r, column=10).value = c.get("clarification", "")
    cov.cell(row=22, column=9).value = sum(len(c.get("tc_nos", [])) for c in items)

    for i, e in enumerate(d.get("exploratory", [])[:EXC_MAX - EXC_START + 1]):
        r = EXC_START + i
        cov.cell(row=r, column=2).value = i + 1
        cov.cell(row=r, column=3).value = e.get("category", "")
        cov.cell(row=r, column=4).value = e.get("case", "")
        cov.cell(row=r, column=5).value = e.get("how", "")
        cov.cell(row=r, column=6).value = e.get("risk", "")
        cov.cell(row=r, column=7).value = e.get("expected", "")
        cov.cell(row=r, column=8).value = ", ".join(e.get("tc_nos", [])) or "신규 케이스 추가 권장"
        cov.cell(row=r, column=9).value = e.get("priority", "")
        cov.cell(row=r, column=10).value = e.get("question", "")

    risks = d.get("risks", [])[:RISK_MAX - RISK_START + 1]
    for i, rk in enumerate(risks):
        r = RISK_START + i
        cov.cell(row=r, column=2).value = rk.get("id", f"R{i+1}")
        cov.cell(row=r, column=3).value = rk.get("name", "")
        cov.cell(row=r, column=4).value = rk.get("reason", "")
        cov.cell(row=r, column=5).value = ", ".join(rk.get("tc_nos", []))
        cov.cell(row=r, column=6).value = len(rk.get("tc_nos", []))
    cov.cell(row=42, column=6).value = sum(len(r.get("tc_nos", [])) for r in risks)

    # ---------- TC_Result_Report ----------
    rep = wb["TC_Result_Report"]
    rep.cell(row=4, column=2).value = total
    rep.cell(row=4, column=12).value = total

    wb.save(out_path)
    return total, len(items)


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else "output/qa-design.json"
    out = sys.argv[2] if len(sys.argv) > 2 else "output/QA_TestCase.xlsx"

    d = json.loads(Path(src).read_text(encoding="utf-8"))
    errors, warnings = validate(d)

    for w in warnings:
        print(f"  [경고] {w}")
    if errors:
        print(f"\n검증 실패 {len(errors)}건:")
        for e in errors:
            print(f"  [오류] {e}")
        print("\n위 항목을 수정한 뒤 다시 실행하세요.")
        return 1

    if not Path(TEMPLATE).exists():
        print(f"[오류] 템플릿을 찾을 수 없음: {TEMPLATE}")
        return 1

    n_tc, n_cov = fill(d, out)
    print(f"TC {n_tc}건 / 커버리지 항목 {n_cov}건 -> {out} 생성 완료")
    return 0


if __name__ == "__main__":
    sys.exit(main())